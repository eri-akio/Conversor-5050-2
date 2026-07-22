"""Leitura e validacao estrutural da planilha (Fase 2).

Ver docs/plano_conversor_dro_5050_simples.md secao 7.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models import (
    CampoNormalizado,
    ETAPA_ESTRUTURA,
    Ocorrencia,
    TIPO_ERRO_IMPEDITIVO,
)
from src.normalizers import (
    normalizar_cnpj,
    normalizar_data_base,
    normalizar_maiusculo,
    normalizar_texto,
)

# Colunas do Cabecalho cujo valor valido e convertido para maiusculo
# (secao 7: codigo/dominio insensivel a caixa).
CABECALHO_COLUNAS_MAIUSCULAS: tuple[str, ...] = (
    "codigoConglomerado",
    "tipoRemessa",
    "opcaoPorProvisaoAcumulada",
)

NOME_ABA_BASE = "Base"
NOME_ABA_CABECALHO = "Cabecalho"

CABECALHO_COLUNAS: tuple[str, ...] = (
    "codigoDocumento",
    "dataBase",
    "codigoConglomerado",
    "cnpj",
    "tipoRemessa",
    "opcaoPorProvisaoAcumulada",
)

BASE_COLUNAS: tuple[str, ...] = (
    "idEvento",
    "categoriaNivel1",
    "categoriaNivel2",
    "tipoAvaliacao",
    "unidadeNegocio",
    "dataDescoberta",
    "dataOcorrencia",
    "naturezaContingencia",
    "codSistemaOrigem",
    "nomeSistema",
    "codigoEventoOrigem",
    "descricaoEvento",
    "riscoAssociado",
    "ligadoRiscoSocioAmbiental",
    "ligadoRiscoCibernetico",
    "negocioDescontinuado",
    "idBacen",
    "probabilidadePerda",
    "valorRisco",
    "dataContabilizacao",
    "contaBalAnaliticoDebito",
    "nomeContaDebito",
    "contaBalAnaliticoCredito",
    "nomeContaCredito",
    "contaCosifDebito",
    "contaCosifCredito",
    "valorPerdaEfetiva",
    "valorProvisao",
    "valorRecuperacao",
    "fonteRecuperacao",
)

METADADO_OPCIONAL = "Source.Name"

ALIAS_CANONICO_POR_NOME_ANTIGO = {
    "ligacaoRiscoSocioambiental": "ligadoRiscoSocioAmbiental",
}


class ArquivoInvalido(Exception):
    """Falha tecnica ao abrir o arquivo (XLSX-001)."""


class PlanilhaInvalida(Exception):
    """Erro impeditivo de estrutura da planilha."""

    def __init__(self, ocorrencia: Ocorrencia) -> None:
        super().__init__(ocorrencia.detalhe)
        self.ocorrencia = ocorrencia


class PlanilhaLida:
    """Abas Base e Cabecalho ja localizadas, prontas para normalizacao."""

    def __init__(
        self,
        aba_base: Worksheet,
        aba_cabecalho: Worksheet,
        cabecalhos_base: tuple[str, ...],
        cabecalhos_cabecalho: tuple[str, ...],
    ) -> None:
        self.aba_base = aba_base
        self.aba_cabecalho = aba_cabecalho
        self.cabecalhos_base = cabecalhos_base
        self.cabecalhos_cabecalho = cabecalhos_cabecalho

    def linhas_base(self) -> list[tuple[object, ...]]:
        return [
            linha
            for linha in self.aba_base.iter_rows(min_row=2, values_only=True)
            if not _linha_totalmente_vazia(linha)
        ]

    def linha_cabecalho(self) -> tuple[object, ...] | None:
        linhas = [
            linha
            for linha in self.aba_cabecalho.iter_rows(min_row=2, values_only=True)
            if not _linha_totalmente_vazia(linha)
        ]
        return linhas[0] if linhas else None


def _linha_totalmente_vazia(linha: tuple[object, ...]) -> bool:
    return all(valor is None or str(valor).strip() == "" for valor in linha)


def abrir_planilha(caminho: Path) -> Workbook:
    if not caminho.exists():
        raise ArquivoInvalido(f"Arquivo nao encontrado: {caminho}")
    if caminho.suffix.lower() != ".xlsx":
        raise ArquivoInvalido(
            f"Extensao invalida (esperado .xlsx): {caminho.suffix}"
        )
    try:
        return load_workbook(caminho, data_only=True, read_only=False)
    except Exception as erro:
        raise ArquivoInvalido(
            f"Falha ao abrir ou ler o arquivo: {erro}"
        ) from erro


def _localizar_aba(workbook: Workbook, nome_esperado: str) -> Worksheet | None:
    alvo = nome_esperado.strip().lower()
    for nome in workbook.sheetnames:
        if nome.strip().lower() == alvo:
            return workbook[nome]
    return None


def _cabecalhos_da_linha_1(aba: Worksheet) -> tuple[str, ...]:
    primeira_linha = next(
        aba.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    return tuple(
        str(valor).strip() if valor is not None else ""
        for valor in primeira_linha
    )


def _validar_colunas_obrigatorias(
    cabecalhos: tuple[str, ...],
    colunas_obrigatorias: tuple[str, ...],
    nome_aba: str,
) -> None:
    presentes = {c for c in cabecalhos if c}

    duplicados = sorted(
        {c for c in cabecalhos if c and cabecalhos.count(c) > 1}
    )
    if duplicados:
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-COL-002",
                descricao="Cabecalho duplicado ou ambiguo.",
                detalhe=(
                    f"Colunas repetidas na aba {nome_aba}: "
                    f"{', '.join(duplicados)}."
                ),
                campos=tuple(duplicados),
            )
        )

    if (
        "ligadoRiscoSocioAmbiental" in presentes
        and "ligacaoRiscoSocioambiental" in presentes
    ):
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-COL-002",
                descricao="Cabecalho duplicado ou ambiguo.",
                detalhe=(
                    "ligadoRiscoSocioAmbiental e o alias "
                    "ligacaoRiscoSocioambiental nao podem coexistir."
                ),
                campos=(
                    "ligadoRiscoSocioAmbiental",
                    "ligacaoRiscoSocioambiental",
                ),
            )
        )

    faltando = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in presentes
        and not (
            coluna == "ligadoRiscoSocioAmbiental"
            and "ligacaoRiscoSocioambiental" in presentes
        )
    ]
    if faltando:
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-COL-001",
                descricao="Coluna obrigatoria ausente.",
                detalhe=(
                    f"Colunas ausentes na aba {nome_aba}: "
                    f"{', '.join(faltando)}."
                ),
                campos=tuple(faltando),
            )
        )


def ler_planilha(caminho: Path) -> PlanilhaLida:
    workbook = abrir_planilha(caminho)

    aba_base = _localizar_aba(workbook, NOME_ABA_BASE)
    aba_cabecalho = _localizar_aba(workbook, NOME_ABA_CABECALHO)

    ausentes = [
        nome
        for nome, aba in (
            (NOME_ABA_BASE, aba_base),
            (NOME_ABA_CABECALHO, aba_cabecalho),
        )
        if aba is None
    ]
    if ausentes:
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-ABA-001",
                descricao="Aba obrigatoria ausente.",
                detalhe=f"Abas ausentes: {', '.join(ausentes)}.",
            )
        )
    assert aba_base is not None
    assert aba_cabecalho is not None

    cabecalhos_base = _cabecalhos_da_linha_1(aba_base)
    cabecalhos_cabecalho = _cabecalhos_da_linha_1(aba_cabecalho)

    _validar_colunas_obrigatorias(
        cabecalhos_base, BASE_COLUNAS, NOME_ABA_BASE
    )
    _validar_colunas_obrigatorias(
        cabecalhos_cabecalho, CABECALHO_COLUNAS, NOME_ABA_CABECALHO
    )

    planilha = PlanilhaLida(
        aba_base=aba_base,
        aba_cabecalho=aba_cabecalho,
        cabecalhos_base=cabecalhos_base,
        cabecalhos_cabecalho=cabecalhos_cabecalho,
    )

    if not planilha.linhas_base():
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-BASE-001",
                descricao="Aba Base sem nenhuma linha de dados.",
                detalhe="A aba Base nao possui nenhuma linha de dados.",
            )
        )

    linhas_cabecalho = [
        linha
        for linha in aba_cabecalho.iter_rows(min_row=2, values_only=True)
        if not _linha_totalmente_vazia(linha)
    ]
    if len(linhas_cabecalho) != 1:
        raise PlanilhaInvalida(
            Ocorrencia(
                etapa=ETAPA_ESTRUTURA,
                tipo=TIPO_ERRO_IMPEDITIVO,
                codigo="XLSX-CAB-002",
                descricao=(
                    "Aba Cabecalho sem nenhuma linha, ou com mais de uma "
                    "linha de dados."
                ),
                detalhe=(
                    f"A aba Cabecalho possui {len(linhas_cabecalho)} "
                    "linha(s) de dados; esperada exatamente 1."
                ),
            )
        )

    return planilha


def extrair_cabecalho(planilha: PlanilhaLida) -> dict[str, CampoNormalizado]:
    """Normaliza a unica linha de dados da aba Cabecalho (Fase 7)."""

    valores = planilha.linha_cabecalho()
    assert valores is not None

    brutos = dict(zip(planilha.cabecalhos_cabecalho, valores))

    campos: dict[str, CampoNormalizado] = {}
    for nome in CABECALHO_COLUNAS:
        valor_bruto = brutos.get(nome)
        if nome == "dataBase":
            campos[nome] = normalizar_data_base(nome, valor_bruto)
        elif nome == "cnpj":
            campos[nome] = normalizar_cnpj(nome, valor_bruto)
        elif nome in CABECALHO_COLUNAS_MAIUSCULAS:
            campos[nome] = normalizar_maiusculo(nome, valor_bruto)
        else:
            campos[nome] = normalizar_texto(nome, valor_bruto)
    return campos
