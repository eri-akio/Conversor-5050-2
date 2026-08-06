"""Relatorio XLSX de duas abas: Resumo e Inconsistencias (Fase 8).

Ver docs/plano_conversor_dro_5050_simples.md secao 21. A aba Inconsistencias
mostra somente problemas: regras aprovadas nao geram nenhuma linha, porque
as funcoes de critica (rules_local/rules_pre/rule_pos) so retornam uma Ocorrencia
quando ha um problema.

Formatacao espelhada do projeto anterior
(src/reporters/xlsx_reporter.py): titulo, cabecalhos de secao, cores de
destaque por status/gravidade, tabela com faixas zebradas e cabecalho
congelado. A cor reforca o texto, nunca o substitui.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from src.models import (
    ETAPA_XSD,
    Ocorrencia,
    TIPO_AVISO,
    TIPO_ERRO_IMPEDITIVO,
    TIPO_FALHA_TECNICA,
)

TITLE_FILL = "17365D"
SECTION_FILL = "1F4E78"
LABEL_FILL = "D9EAF7"
HEADER_FONT_COLOR = "FFFFFF"
BORDER_COLOR = "B7C9D6"

REGRAS_NAO_EXECUTADAS = (
    "DRO001002",
    "DRO000016",
    "DRO000017",
    "DRO000022",
    "DRO000026",
    "DRO000027",
    "DRO000028",
    "DRO000029",
    "DRO000030",
)

STATUS_FILLS = {
    "APROVADO": "C6EFCE",
    "REPROVADO": "FFC7CE",
    "NÃO EXECUTADO": "D9D9D9",
    "FALHA TÉCNICA": "F4CCCC",
}

CABECALHOS_INCONSISTENCIAS: tuple[str, ...] = (
    "Etapa",
    "Tipo",
    "Linha(s) da planilha",
    "idEvento",
    "Campo(s)",
    "Código da regra",
    "Descrição da regra",
    "Detalhe da inconsistência",
)

LARGURAS_INCONSISTENCIAS: tuple[int, ...] = (18, 16, 14, 12, 22, 16, 42, 60)


def _fill(cor: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=cor)


def _texto_linhas(ocorrencia: Ocorrencia) -> str:
    return ", ".join(str(numero) for numero in ocorrencia.linhas)


def _texto_campos(ocorrencia: Ocorrencia) -> str:
    return ", ".join(ocorrencia.campos)


def gerar_relatorio(
    caminho: Path,
    *,
    status_local: str,
    status_xsd: str,
    ocorrencias: list[Ocorrencia],
) -> None:
    if caminho.exists():
        raise FileExistsError(
            f"Arquivo já existe, não será sobrescrito: {caminho}"
        )

    workbook = Workbook()

    aba_resumo = workbook.active
    aba_resumo.title = "Resumo"
    _escrever_resumo(aba_resumo, status_local, status_xsd, ocorrencias)

    aba_inconsistencias = workbook.create_sheet("Inconsistencias")
    _escrever_inconsistencias(aba_inconsistencias, ocorrencias)

    workbook.save(caminho)


def _titulo_secao(aba: Worksheet, intervalo: str, texto: str) -> None:
    aba.merge_cells(intervalo)
    celula = aba[intervalo.split(":", maxsplit=1)[0]]
    celula.value = texto
    celula.alignment = Alignment(horizontal="center", vertical="center")
    for linha in aba[intervalo]:
        for item in linha:
            item.fill = _fill(SECTION_FILL)
            item.font = Font(bold=True, color=HEADER_FONT_COLOR)


def _escrever_resumo(
    aba: Worksheet,
    status_local: str,
    status_xsd: str,
    ocorrencias: list[Ocorrencia],
) -> None:
    aba.column_dimensions["A"].width = 30
    aba.column_dimensions["B"].width = 18

    aba.merge_cells("A1:B2")
    titulo = aba["A1"]
    titulo.value = "RELATÓRIO DE EXECUÇÃO — DRO 5050"
    titulo.font = Font(bold=True, color=HEADER_FONT_COLOR, size=14)
    titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for linha in aba["A1:B2"]:
        for item in linha:
            item.fill = _fill(TITLE_FILL)

    _titulo_secao(aba, "A3:B3", "RESULTADO DA VALIDAÇÃO")

    borda_inferior = Side(style="thin", color=BORDER_COLOR)
    linhas_resultado = (
        ("Validação local", status_local),
        ("Validação XSD", status_xsd),
    )
    for indice, (rotulo, valor) in enumerate(linhas_resultado, start=4):
        celula_rotulo = aba.cell(row=indice, column=1, value=rotulo)
        celula_rotulo.fill = _fill(LABEL_FILL)
        celula_rotulo.font = Font(bold=True)
        celula_rotulo.border = Border(bottom=borda_inferior)

        celula_valor = aba.cell(row=indice, column=2, value=valor)
        celula_valor.border = Border(bottom=borda_inferior)
        cor = STATUS_FILLS.get(valor)
        if cor is not None:
            celula_valor.fill = _fill(cor)
            celula_valor.font = Font(bold=True)

    linha_indicadores = 4 + len(linhas_resultado) + 1
    _titulo_secao(
        aba,
        f"A{linha_indicadores}:B{linha_indicadores}",
        "INDICADORES DA EXECUÇÃO",
    )

    indicadores = (
        ("Total de inconsistências", len(ocorrencias)),
        (
            "Regras com inconsistência",
            len({o.codigo for o in ocorrencias}),
        ),
        (
            "Eventos com inconsistência",
            len(
                {
                    o.id_evento
                    for o in ocorrencias
                    if o.id_evento is not None
                }
            ),
        ),
        (
            "Erros impeditivos",
            sum(1 for o in ocorrencias if o.tipo == TIPO_ERRO_IMPEDITIVO),
        ),
        ("Avisos", sum(1 for o in ocorrencias if o.tipo == TIPO_AVISO)),
        (
            "Falhas técnicas",
            sum(1 for o in ocorrencias if o.tipo == TIPO_FALHA_TECNICA),
        ),
        (
            "Erros XSD",
            sum(1 for o in ocorrencias if o.etapa == ETAPA_XSD),
        ),
        ("Regras não executadas", len(REGRAS_NAO_EXECUTADAS)),
        ("Códigos não executados", ", ".join(REGRAS_NAO_EXECUTADAS)),
    )
    for indice, (rotulo, valor) in enumerate(
        indicadores, start=linha_indicadores + 1
    ):
        celula_rotulo = aba.cell(row=indice, column=1, value=rotulo)
        celula_rotulo.fill = _fill(LABEL_FILL)
        celula_rotulo.font = Font(bold=True)
        celula_rotulo.border = Border(bottom=borda_inferior)

        aba.cell(row=indice, column=2, value=valor).border = Border(
            bottom=borda_inferior
        )

    aba.freeze_panes = "A3"


def _escrever_inconsistencias(
    aba: Worksheet, ocorrencias: list[Ocorrencia]
) -> None:
    aba.append(CABECALHOS_INCONSISTENCIAS)
    for ocorrencia in ocorrencias:
        aba.append(
            (
                ocorrencia.etapa,
                ocorrencia.tipo,
                _texto_linhas(ocorrencia),
                ocorrencia.id_evento or "",
                _texto_campos(ocorrencia),
                ocorrencia.codigo,
                ocorrencia.descricao,
                ocorrencia.detalhe,
            )
        )

    linha_final = aba.max_row
    coluna_final = len(CABECALHOS_INCONSISTENCIAS)

    for indice, largura in enumerate(LARGURAS_INCONSISTENCIAS, start=1):
        letra = get_column_letter(indice)
        aba.column_dimensions[letra].width = largura

    for celula in aba[1]:
        celula.fill = _fill(SECTION_FILL)
        celula.font = Font(bold=True, color=HEADER_FONT_COLOR)
        celula.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for linha in aba.iter_rows(min_row=2, max_row=linha_final):
        for indice, celula in enumerate(linha, start=1):
            celula.alignment = Alignment(
                vertical="top",
                wrap_text=indice in (7, 8),
            )

    if ocorrencias:
        tabela = Table(
            displayName="Inconsistencias",
            ref=f"A1:{get_column_letter(coluna_final)}{linha_final}",
        )
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        aba.add_table(tabela)

        intervalo_tipo = f"B2:B{linha_final}"
        aba.conditional_formatting.add(
            intervalo_tipo,
            FormulaRule(
                formula=[f'$B2="{TIPO_ERRO_IMPEDITIVO}"'],
                fill=_fill("F4CCCC"),
                font=Font(bold=True, color="9C0006"),
            ),
        )
        aba.conditional_formatting.add(
            intervalo_tipo,
            FormulaRule(
                formula=[f'$B2="{TIPO_AVISO}"'],
                fill=_fill("FFF2CC"),
                font=Font(color="7F6000"),
            ),
        )
    else:
        aba.auto_filter.ref = aba.dimensions

    aba.freeze_panes = "A2"
