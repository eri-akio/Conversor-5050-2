"""Testes da Fase 8: relatorio XLSX de duas abas
(src/report_writer.py)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import (
    ETAPA_GERACAO_XML,
    ETAPA_PRE_PROCESSAMENTO,
    ETAPA_XSD,
    Ocorrencia,
    TIPO_AVISO,
    TIPO_ERRO_IMPEDITIVO,
    TIPO_FALHA_TECNICA,
)
from src.report_writer import gerar_relatorio


def _ocorrencia(**sobrescritas: object) -> Ocorrencia:
    base = dict(
        etapa=ETAPA_PRE_PROCESSAMENTO,
        tipo=TIPO_ERRO_IMPEDITIVO,
        codigo="DRO001231",
        descricao="Descrição de exemplo.",
        detalhe="Detalhe de exemplo.",
        linhas=(2,),
        id_evento="EVT-1",
        campos=("valorPerdaEfetiva",),
    )
    base.update(sobrescritas)
    return Ocorrencia(**base)


def test_relatorio_tem_exatamente_duas_abas(tmp_path: Path) -> None:
    caminho = tmp_path / "relatorio.xlsx"

    gerar_relatorio(
        caminho,
        status_local="REPROVADO",
        status_xsd="NÃO EXECUTADO",
        ocorrencias=[_ocorrencia()],
    )

    workbook = load_workbook(caminho)

    assert workbook.sheetnames == ["Resumo", "Inconsistencias"]


def test_aba_resumo_contem_status_e_indicadores(tmp_path: Path) -> None:
    caminho = tmp_path / "relatorio.xlsx"
    ocorrencias = [
        _ocorrencia(codigo="DRO001231", id_evento="EVT-1"),
        _ocorrencia(
            codigo="DRO000024", tipo=TIPO_AVISO, id_evento="EVT-2"
        ),
        _ocorrencia(
            codigo="XSD-001",
            etapa=ETAPA_XSD,
            id_evento=None,
            linhas=(),
        ),
    ]

    gerar_relatorio(
        caminho,
        status_local="REPROVADO",
        status_xsd="REPROVADO",
        ocorrencias=ocorrencias,
    )

    aba = load_workbook(caminho)["Resumo"]
    valores = {linha[0]: linha[1] for linha in aba.iter_rows(values_only=True) if linha and linha[0]}

    assert valores["Validação local"] == "REPROVADO"
    assert valores["Validação XSD"] == "REPROVADO"
    assert valores["Total de inconsistências"] == 3
    assert valores["Regras com inconsistência"] == 3
    assert valores["Eventos com inconsistência"] == 2
    assert valores["Erros impeditivos"] == 2
    assert valores["Avisos"] == 1
    assert valores["Falhas t\u00e9cnicas"] == 0
    assert valores["Erros XSD"] == 1
    assert valores["Regras não executadas"] == 9
    codigos_nao_executados = valores["Códigos não executados"]
    assert "DRO001002" in codigos_nao_executados
    assert "DRO000030" in codigos_nao_executados


def test_aba_inconsistencias_tem_cabecalho_e_uma_linha_por_ocorrencia(
    tmp_path: Path,
) -> None:
    caminho = tmp_path / "relatorio.xlsx"
    ocorrencias = [_ocorrencia(), _ocorrencia(codigo="DRO001232")]

    gerar_relatorio(
        caminho,
        status_local="REPROVADO",
        status_xsd="NÃO EXECUTADO",
        ocorrencias=ocorrencias,
    )

    aba = load_workbook(caminho)["Inconsistencias"]
    linhas = list(aba.iter_rows(values_only=True))

    assert linhas[0] == (
        "Etapa",
        "Tipo",
        "Linha(s) da planilha",
        "idEvento",
        "Campo(s)",
        "Código da regra",
        "Descrição da regra",
        "Detalhe da inconsistência",
    )
    assert len(linhas) == 3
    assert linhas[1][5] == "DRO001231"
    assert linhas[1][2] == "2"
    assert linhas[2][5] == "DRO001232"


def test_relatorio_sem_ocorrencias_mostra_zero_e_aba_so_com_cabecalho(
    tmp_path: Path,
) -> None:
    caminho = tmp_path / "relatorio.xlsx"

    gerar_relatorio(
        caminho,
        status_local="APROVADO",
        status_xsd="APROVADO",
        ocorrencias=[],
    )

    workbook = load_workbook(caminho)
    aba_resumo = workbook["Resumo"]
    valores = {
        linha[0]: linha[1]
        for linha in aba_resumo.iter_rows(values_only=True)
        if linha and linha[0]
    }
    assert valores["Total de inconsistências"] == 0

    linhas_inconsistencias = list(
        workbook["Inconsistencias"].iter_rows(values_only=True)
    )
    assert len(linhas_inconsistencias) == 1  # so o cabecalho


def test_titulo_e_status_tem_formatacao_visual(tmp_path: Path) -> None:
    caminho = tmp_path / "relatorio.xlsx"

    gerar_relatorio(
        caminho,
        status_local="REPROVADO",
        status_xsd="APROVADO",
        ocorrencias=[_ocorrencia()],
    )

    aba = load_workbook(caminho)["Resumo"]

    assert aba["A1"].value == "RELATÓRIO DE EXECUÇÃO — DRO 5050"
    assert aba["A1"].font.bold is True
    assert aba["A1"].fill.fgColor.rgb == "0017365D"

    # "Validação local" fica na linha 4, coluna B (ver _escrever_resumo).
    celula_status_local = aba.cell(row=4, column=2)
    assert celula_status_local.value == "REPROVADO"
    assert celula_status_local.fill.fgColor.rgb == "00FFC7CE"

    celula_status_xsd = aba.cell(row=5, column=2)
    assert celula_status_xsd.value == "APROVADO"
    assert celula_status_xsd.fill.fgColor.rgb == "00C6EFCE"


def test_inconsistencias_vira_tabela_com_faixas_e_formatacao_condicional(
    tmp_path: Path,
) -> None:
    caminho = tmp_path / "relatorio.xlsx"
    ocorrencias = [_ocorrencia(), _ocorrencia(codigo="DRO000024", tipo=TIPO_AVISO)]

    gerar_relatorio(
        caminho,
        status_local="REPROVADO",
        status_xsd="NÃO EXECUTADO",
        ocorrencias=ocorrencias,
    )

    aba = load_workbook(caminho)["Inconsistencias"]

    assert "Inconsistencias" in aba.tables
    assert aba.freeze_panes == "A2"
    assert aba["A1"].font.bold is True
    assert aba["A1"].fill.fgColor.rgb == "001F4E78"
    assert len(aba.conditional_formatting._cf_rules) >= 1


def test_relatorio_nao_sobrescreve_arquivo_existente(tmp_path: Path) -> None:
    caminho = tmp_path / "relatorio.xlsx"
    caminho.write_bytes(b"conteudo previo")

    with pytest.raises(FileExistsError):
        gerar_relatorio(
            caminho,
            status_local="APROVADO",
            status_xsd="APROVADO",
            ocorrencias=[],
        )

    assert caminho.read_bytes() == b"conteudo previo"

def test_falha_de_geracao_xml_nao_e_contada_como_erro_xsd(
    tmp_path: Path,
) -> None:
    caminho = tmp_path / "relatorio.xlsx"
    gerar_relatorio(
        caminho,
        status_local="APROVADO",
        status_xsd="FALHA T\u00c9CNICA",
        ocorrencias=[
            _ocorrencia(
                codigo="XML-TEC-001",
                etapa=ETAPA_GERACAO_XML,
                tipo=TIPO_FALHA_TECNICA,
                id_evento=None,
                linhas=(),
            )
        ],
    )

    aba = load_workbook(caminho)["Resumo"]
    valores = {
        linha[0]: linha[1]
        for linha in aba.iter_rows(values_only=True)
        if linha and linha[0]
    }
    assert valores["Falhas t\u00e9cnicas"] == 1
    assert valores["Erros impeditivos"] == 0
    assert valores["Erros XSD"] == 0
