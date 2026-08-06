"""Testes da Fase 9: orquestracao do fluxo completo
(src/conversion.py)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

import src.conversion as conversion
import src.xsd_validator as xsd_validator
from src.conversion import processar
from src.reader import BASE_COLUNAS, CABECALHO_COLUNAS

CABECALHO_VALIDO = {
    "codigoDocumento": "5050",
    "dataBase": "2025-06",
    "codigoConglomerado": "C0099999",
    "cnpj": "12345678",
    "tipoRemessa": "I",
    "opcaoPorProvisaoAcumulada": "S",
}

CAMPOS_EVENTO_PADRAO = {
    "idEvento": "EVT1",
    "categoriaNivel1": "1",
    "categoriaNivel2": "11",
    "tipoAvaliacao": "NA",
    "unidadeNegocio": "1",
    "dataOcorrencia": "2025-06-10",
    "dataDescoberta": "2025-06-10",
    "naturezaContingencia": "NA",
    "codSistemaOrigem": "SIS1",
    "nomeSistema": "Sistema Um",
    "codigoEventoOrigem": "COD1",
    "riscoAssociado": "NA",
    "ligadoRiscoSocioAmbiental": "N",
    "ligadoRiscoCibernetico": "N",
    "idBacen": "Z0000001",
    "dataContabilizacao": "2025-06-15",
    "valorPerdaEfetiva": "1000.00",
    "valorProvisao": 0,
    "valorRecuperacao": 0,
    "contaBalAnaliticoDebito": "123456",
    "nomeContaDebito": "Conta Debito",
    "contaCosifDebito": "10000007",
    "contaBalAnaliticoCredito": "654321",
    "nomeContaCredito": "Conta Credito",
    "contaCosifCredito": "20000006",
}


def _construir_planilha_valida(tmp_path: Path) -> Path:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    aba_base.append(
        [CAMPOS_EVENTO_PADRAO.get(coluna) for coluna in BASE_COLUNAS]
    )
    # Segundo evento, abaixo dos limiares de individualizacao: garante que
    # eventosConsolidados tenha ao menos 1 elemento (minOccurs=1 no XSD).
    campos_consolidavel = dict(CAMPOS_EVENTO_PADRAO)
    campos_consolidavel.update(
        idEvento="EVT2",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD2",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append(
        [campos_consolidavel.get(coluna) for coluna in BASE_COLUNAS]
    )

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )

    caminho = tmp_path / "planilha.xlsx"
    workbook.save(caminho)
    return caminho


def _construir_planilha_valida_com_evento(
    tmp_path: Path, **sobrescritas: object
) -> Path:
    """Igual a _construir_planilha_valida, mas permite sobrescrever campos
    do primeiro evento (individualizado) -- usada para cenarios APROVADO
    com XML gerado, quando o teste precisa customizar algum campo."""

    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    campos = dict(CAMPOS_EVENTO_PADRAO)
    campos.update(sobrescritas)
    aba_base.append([campos.get(coluna) for coluna in BASE_COLUNAS])

    campos_consolidavel = dict(CAMPOS_EVENTO_PADRAO)
    campos_consolidavel.update(
        idEvento="EVT2",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD2",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append(
        [campos_consolidavel.get(coluna) for coluna in BASE_COLUNAS]
    )

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )

    caminho = tmp_path / "planilha.xlsx"
    workbook.save(caminho)
    return caminho


def _construir_planilha_com_evento(tmp_path: Path, **sobrescritas: object) -> Path:
    """Planilha com um unico evento na aba Base, para cenarios REPROVADO
    (nao precisa de segundo evento consolidavel — XML nao chega a ser
    construido quando status_local != APROVADO)."""

    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    campos = dict(CAMPOS_EVENTO_PADRAO)
    campos.update(sobrescritas)
    aba_base.append([campos.get(coluna) for coluna in BASE_COLUNAS])

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "planilha.xlsx"
    workbook.save(caminho_planilha)
    return caminho_planilha


def test_planilha_valida_produz_xml_aprovado_e_relatorio(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_valida(tmp_path)
    pasta_saida = tmp_path / "saida"

    resultado = processar(caminho_planilha, pasta_saida)

    assert resultado.status_local == "APROVADO"
    assert resultado.status_xsd == "APROVADO"
    assert resultado.caminho_xml is not None
    assert resultado.caminho_xml.exists()
    assert resultado.caminho_relatorio is not None
    assert resultado.caminho_relatorio.exists()
    assert resultado.caminho_xml.name == "DRO_5050_2025-06.xml"
    assert resultado.caminho_relatorio.name == "Relatorio_DRO_5050_2025-06.xlsx"


def test_arquivo_inexistente_gera_falha_tecnica_sem_relatorio(
    tmp_path: Path,
) -> None:
    resultado = processar(tmp_path / "nao_existe.xlsx", tmp_path / "saida")

    assert resultado.status_local == "FALHA TÉCNICA"
    assert resultado.status_xsd == "NÃO EXECUTADO"
    assert resultado.caminho_xml is None
    assert resultado.caminho_relatorio is None


def test_aba_ausente_gera_relatorio_reprovado_sem_xml(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Base"
    workbook.active.append(list(BASE_COLUNAS))
    workbook.active.append(
        [CAMPOS_EVENTO_PADRAO.get(coluna) for coluna in BASE_COLUNAS]
    )
    caminho_planilha = tmp_path / "sem_cabecalho.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.status_xsd == "NÃO EXECUTADO"
    assert resultado.caminho_xml is None
    assert resultado.caminho_relatorio is not None
    assert resultado.caminho_relatorio.exists()
    assert any(o.codigo == "XLSX-ABA-001" for o in resultado.ocorrencias)


def test_saldo_negativo_reprova_pelas_criticas_oficiais(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    campos_invalidos = dict(CAMPOS_EVENTO_PADRAO)
    campos_invalidos["valorPerdaEfetiva"] = "-50"
    aba_base.append(
        [campos_invalidos.get(coluna) for coluna in BASE_COLUNAS]
    )

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "com_erro.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.status_xsd == "NÃO EXECUTADO"
    assert resultado.caminho_xml is None
    codigos = [o.codigo for o in resultado.ocorrencias]
    assert "BASE-SINAL-CONT-001" not in codigos
    assert "DRO000011" in codigos
    assert "DRO000023" in codigos


def test_conta_cosif_fora_do_cadastro_reprova_e_nao_gera_xml(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    campos_invalidos = dict(CAMPOS_EVENTO_PADRAO)
    campos_invalidos["contaCosifDebito"] = "99999999"  # fora do cadastro COSIF
    aba_base.append(
        [campos_invalidos.get(coluna) for coluna in BASE_COLUNAS]
    )

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "cosif_invalido.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.status_xsd == "NÃO EXECUTADO"
    assert resultado.caminho_xml is None
    assert any(o.codigo == "DRO001431" for o in resultado.ocorrencias)


def test_segunda_execucao_na_mesma_pasta_gera_arquivos_com_sufixo(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_valida(tmp_path)
    pasta_saida = tmp_path / "saida"

    primeiro = processar(caminho_planilha, pasta_saida)
    segundo = processar(caminho_planilha, pasta_saida)

    assert primeiro.caminho_xml.name == "DRO_5050_2025-06.xml"
    assert primeiro.caminho_relatorio.name == "Relatorio_DRO_5050_2025-06.xlsx"
    assert segundo.caminho_xml.name == "DRO_5050_2025-06_1.xml"
    assert (
        segundo.caminho_relatorio.name
        == "Relatorio_DRO_5050_2025-06_1.xlsx"
    )
    assert primeiro.caminho_xml.exists()
    assert segundo.caminho_xml.exists()

    terceiro = processar(caminho_planilha, pasta_saida)
    assert terceiro.caminho_xml.name == "DRO_5050_2025-06_2.xml"


def _construir_planilha_com_cabecalho(
    tmp_path: Path, **sobrescritas_cabecalho: object
) -> Path:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    aba_base.append(
        [CAMPOS_EVENTO_PADRAO.get(coluna) for coluna in BASE_COLUNAS]
    )

    cabecalho = dict(CABECALHO_VALIDO)
    cabecalho.update(sobrescritas_cabecalho)
    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append([cabecalho.get(coluna) for coluna in CABECALHO_COLUNAS])

    caminho = tmp_path / "planilha.xlsx"
    workbook.save(caminho)
    return caminho


def test_data_base_ausente_reprova_sem_crash_e_gera_relatorio(
    tmp_path: Path,
) -> None:
    """P1: dataBase ausente nao pode derrubar o processo com
    ValueError/year 0 is out of range — deve reprovar normalmente."""
    caminho_planilha = _construir_planilha_com_cabecalho(
        tmp_path, dataBase=None
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.caminho_xml is None
    assert resultado.caminho_relatorio is not None
    assert resultado.caminho_relatorio.exists()
    assert resultado.caminho_relatorio.name == (
        "Relatorio_DRO_5050_SEM_DATA_BASE.xlsx"
    )
    assert any(o.codigo == "BASE-CAB-DATABASE-001" for o in resultado.ocorrencias)


def test_data_base_com_mes_invalido_reprova_localmente(
    tmp_path: Path,
) -> None:
    """P2: dataBase="2026-07" (mes fora de {06,12}) e reprovada."""
    caminho_planilha = _construir_planilha_com_cabecalho(
        tmp_path, dataBase="2026-07"
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.caminho_xml is None
    assert any(o.codigo == "BASE-CAB-DATABASE-001" for o in resultado.ocorrencias)


def test_id_evento_colidindo_reprova_com_ocorrencia_de_colisao(
    tmp_path: Path,
) -> None:
    """P8: 'IND-0001' e 'IND0001' colidem no mesmo idEvento normalizado."""
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    linha_1 = dict(CAMPOS_EVENTO_PADRAO)
    linha_1["idEvento"] = "IND-0001"
    aba_base.append([linha_1.get(coluna) for coluna in BASE_COLUNAS])
    linha_2 = dict(CAMPOS_EVENTO_PADRAO)
    linha_2["idEvento"] = "IND0001"
    linha_2["codigoEventoOrigem"] = "COD2"
    aba_base.append([linha_2.get(coluna) for coluna in BASE_COLUNAS])

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "colisao.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.caminho_xml is None
    assert any(
        o.codigo == "BASE-IDEVENTO-COLISAO-001" for o in resultado.ocorrencias
    )


def test_xsd_indisponivel_nao_derruba_o_processo(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """#5: XSD ausente/corrompido nao pode propagar excecao -- a
    validacao local ja tinha terminado com sucesso, entao status_local
    permanece APROVADO e so status_xsd vira FALHA TÉCNICA."""

    monkeypatch.setattr(
        xsd_validator, "XSD_PATH", tmp_path / "nao_existe.xsd"
    )
    caminho_planilha = _construir_planilha_valida(tmp_path)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "APROVADO"
    assert resultado.status_xsd == "FALHA TÉCNICA"
    assert resultado.caminho_xml is None
    assert resultado.caminho_relatorio is not None
    assert resultado.caminho_relatorio.exists()
    ocorrencia = next(
        o for o in resultado.ocorrencias if o.codigo == "XSD-TEC-001"
    )
    assert ocorrencia.tipo == "FALHA T\u00c9CNICA"
    assert ocorrencia.etapa == "Valida\u00e7\u00e3o XSD"


# ---------------------------------------------------------------------------
# Formato/dominio ponta a ponta (validar_formatos_e_dominios_evento +
# curto-circuito de conversion.py)
# ---------------------------------------------------------------------------


def test_categoria_nivel2_fora_do_dominio_e_pega_localmente_antes_do_xsd(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path, categoriaNivel2="19"
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.status_xsd == "NÃO EXECUTADO"
    assert resultado.caminho_xml is None
    assert any(
        o.codigo == "BASE-CATEGORIA2-FORM-001" for o in resultado.ocorrencias
    )


def test_avaliacao_invalida_suprime_regra_de_negocio_correspondente(
    tmp_path: Path,
) -> None:
    """tipoAvaliacao='X' e um dominio invalido, nao um estado 'ausente' ou
    'NA' — sem o curto-circuito, BASE-CONT-001 (natureza incompativel com
    avaliacao) tambem dispararia para a mesma causa raiz."""
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path, tipoAvaliacao="X", naturezaContingencia="TRI"
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-AVALIACAO-FORM-001" in codigos
    assert "BASE-CONT-001" not in codigos


@pytest.mark.parametrize("valor", ["IE", "ME"])
def test_tipo_avaliacao_ie_me_e_aceito_e_emitido_sem_conversao_no_xml(
    tmp_path: Path, valor: str
) -> None:
    """Instrucao 12/2026: tipoAvaliacao tambem aceita IE/ME (processos
    encerrados). Por decisao do usuario, sao aceitos incondicionalmente e
    emitidos no XML tal como informados, sem conversao para I/M e sem
    bloquear o status de aprovacao (diverge da politica mais cautelosa
    registrada em CONF-002/VER-001 em docs/conflitos_documentais.md)."""
    caminho_planilha = _construir_planilha_valida_com_evento(
        tmp_path, tipoAvaliacao=valor
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-AVALIACAO-FORM-001" not in codigos
    assert resultado.status_local == "APROVADO"
    assert resultado.status_xsd == "APROVADO"
    assert resultado.caminho_xml is not None
    conteudo_xml = resultado.caminho_xml.read_text(encoding="utf-8")
    assert f'tipoAvaliacao="{valor}"' in conteudo_xml


def test_natureza_contingencia_out_e_aceita_e_emitida_sem_conversao_no_xml(
    tmp_path: Path,
) -> None:
    """Instrucao 12/2026: naturezaContingencia tambem aceita OUT (outras
    contingencias). Por decisao do usuario, mesmo tratamento dado a IE/ME:
    aceito incondicionalmente e emitido no XML tal como informado, sem
    bloquear o status de aprovacao (diverge da politica mais cautelosa
    registrada em CONF-003/VER-001 em docs/conflitos_documentais.md)."""
    caminho_planilha = _construir_planilha_valida_com_evento(
        tmp_path, naturezaContingencia="OUT"
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-NATUREZA-FORM-001" not in codigos
    assert resultado.status_local == "APROVADO"
    assert resultado.status_xsd == "APROVADO"
    assert resultado.caminho_xml is not None
    conteudo_xml = resultado.caminho_xml.read_text(encoding="utf-8")
    assert 'naturezaContingencia="OUT"' in conteudo_xml


# NOTA: validar_natureza_contingencia_avaliacao (BASE-CONT-001) ainda nao e
# chamada por nenhum orquestrador (validar_estrutura_evento/validar_evento_local
# em src/rules_local.py, nem conversion.py) -- e uma regra pronta mas ainda
# nao integrada ao pipeline, algo pre-existente e fora do escopo desta
# mudanca. Por isso, a consequencia de BASE-CONT-001 rejeitar IE/ME com
# naturezaContingencia real e testada no nivel de unidade, diretamente sobre
# a funcao, em tests/test_rules_pre.py
# (test_natureza_tri_com_avaliacao_ie_gera_base_cont_001), nao aqui.


def test_dro001302_dispara_mesmo_com_erro_de_formato_em_campo_nao_relacionado(
    tmp_path: Path,
) -> None:
    """categoriaNivel2 invalido aciona o curto-circuito de formato
    (BASE-CATEGORIA2-FORM-001), que suprime validar_evento() -- mas a
    DRO001302 e chamada separadamente por conversion.py, antes desse
    curto-circuito, e deve continuar aparecendo mesmo assim."""
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path,
        tipoAvaliacao="I",
        naturezaContingencia="TRI",
        categoriaNivel2="19",
        valorProvisao=None,
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert [ocorrencia.codigo for ocorrencia in resultado.ocorrencias] == [
        "BASE-CONT-OBR-001",
        "DRO001302",
        "BASE-CATEGORIA2-FORM-001",
    ]


def test_probabilidade_e_categoria_malformadas_aparecem_juntas(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path,
        categoriaNivel2="19",
        probabilidadePerda="XX",
        valorRisco=100,
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-PROBABILIDADE-FORM-001" in codigos
    assert "BASE-CATEGORIA2-FORM-001" in codigos


def test_probabilidade_invalida_com_avaliacao_individual_suprime_dro001312(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path,
        tipoAvaliacao="I",
        probabilidadePerda="XX",
        valorRisco=100,
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-PROBABILIDADE-FORM-001" in codigos
    assert "DRO001312" not in codigos
    assert resultado.status_local == "REPROVADO"
    assert resultado.status_xsd == "NÃO EXECUTADO"


def test_divergencia_entre_linhas_do_evento_tem_precedencia_sobre_formato(
    tmp_path: Path,
) -> None:
    """Duas linhas do mesmo idEvento com tipoAvaliacao divergente: o evento
    fica inconsistente (BASE-AGR-001) e validar_formatos_e_dominios_evento
    nem chega a rodar nesta execucao (so roda quando evento.consistente)."""
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))
    linha_1 = dict(CAMPOS_EVENTO_PADRAO)
    linha_1["tipoAvaliacao"] = "I"
    aba_base.append([linha_1.get(coluna) for coluna in BASE_COLUNAS])
    linha_2 = dict(CAMPOS_EVENTO_PADRAO)
    linha_2["tipoAvaliacao"] = "XX"
    aba_base.append([linha_2.get(coluna) for coluna in BASE_COLUNAS])

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "divergencia.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    codigos = {o.codigo for o in resultado.ocorrencias}
    assert "BASE-AGR-001" in codigos
    assert "BASE-AVALIACAO-FORM-001" not in codigos


def test_evento_com_erro_de_formato_e_excluido_da_consolidacao_sem_crash(
    tmp_path: Path,
) -> None:
    """Documento com tres eventos: um malformado, um consolidavel e um
    individual. O primeiro nao pode entrar na consolidacao nem bloquear a
    DRO000004 do terceiro; o documento permanece reprovado apenas no escopo
    local correspondente a cada evento."""
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))

    evento_malformado = dict(CAMPOS_EVENTO_PADRAO)
    evento_malformado.update(
        idEvento="EVT1", categoriaNivel2="19", codigoEventoOrigem="COD1"
    )
    aba_base.append(
        [evento_malformado.get(coluna) for coluna in BASE_COLUNAS]
    )

    evento_consolidavel = dict(CAMPOS_EVENTO_PADRAO)
    evento_consolidavel.update(
        idEvento="EVT2",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD2",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append(
        [evento_consolidavel.get(coluna) for coluna in BASE_COLUNAS]
    )

    evento_individual = dict(CAMPOS_EVENTO_PADRAO)
    evento_individual.update(
        idEvento="EVT3",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD3",
        tipoAvaliacao="I",
        naturezaContingencia="TRI",
        probabilidadePerda="PR",
        valorRisco=1000,
        valorPerdaEfetiva="1000.00",
    )
    aba_base.append(
        [evento_individual.get(coluna) for coluna in BASE_COLUNAS]
    )

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "consolidacao_com_erro.xlsx"
    workbook.save(caminho_planilha)

    resultado = processar(caminho_planilha, tmp_path / "saida")

    assert resultado.status_local == "REPROVADO"
    assert resultado.caminho_xml is None
    assert any(
        o.codigo == "BASE-CATEGORIA2-FORM-001" for o in resultado.ocorrencias
    )
    assert any(
        ocorrencia.codigo == "DRO000004"
        and ocorrencia.id_evento == "EVT3"
        for ocorrencia in resultado.ocorrencias
    )

def test_erro_de_normalizacao_bloqueia_regras_e_consolidacao_do_evento(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))

    evento_invalido = dict(CAMPOS_EVENTO_PADRAO)
    evento_invalido.update(
        idEvento="EVT1",
        dataOcorrencia="x",
        codigoEventoOrigem="COD1",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append(
        [evento_invalido.get(coluna) for coluna in BASE_COLUNAS]
    )

    evento_valido = dict(CAMPOS_EVENTO_PADRAO)
    evento_valido.update(
        idEvento="EVT2",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD2",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append([evento_valido.get(coluna) for coluna in BASE_COLUNAS])

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho_planilha = tmp_path / "normalizacao_invalida.xlsx"
    workbook.save(caminho_planilha)

    eventos_consolidados: list[str] = []
    consolidar_original = conversion.consolidar_eventos

    def registrar_consolidacao(eventos, data_base):
        eventos_consolidados.extend(eventos)
        return consolidar_original(eventos, data_base)

    monkeypatch.setattr(conversion, "consolidar_eventos", registrar_consolidacao)

    resultado = conversion.processar(caminho_planilha, tmp_path / "saida")

    assert any(
        ocorrencia.codigo == "BASE-NULO-001"
        and ocorrencia.id_evento == "EVT1"
        for ocorrencia in resultado.ocorrencias
    )
    assert not any(
        ocorrencia.codigo.startswith(("DRO001", "DRO000"))
        and ocorrencia.id_evento == "EVT1"
        for ocorrencia in resultado.ocorrencias
    )
    assert eventos_consolidados == ["EVT2"]

def test_dro001452_permanece_visivel_com_bloco_contabil_incompleto(
    tmp_path: Path,
) -> None:
    caminho_planilha = _construir_planilha_com_evento(
        tmp_path,
        tipoAvaliacao="I",
        naturezaContingencia="TRI",
        probabilidadePerda="PR",
        valorRisco=100,
        dataContabilizacao=None,
        valorPerdaEfetiva=0,
        valorProvisao=0,
        valorRecuperacao=0,
    )

    resultado = processar(caminho_planilha, tmp_path / "saida")
    codigos = [ocorrencia.codigo for ocorrencia in resultado.ocorrencias]

    assert "BASE-CONT-OBR-001" in codigos
    assert "DRO001452" in codigos

def test_erro_oficial_impeditivo_tambem_exclui_evento_da_consolidacao(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(BASE_COLUNAS))

    evento_com_dro = dict(CAMPOS_EVENTO_PADRAO)
    evento_com_dro.update(
        idEvento="EVT1",
        codigoEventoOrigem="COD1",
        valorPerdaEfetiva="10.00",
        valorProvisao=500,
    )
    aba_base.append([evento_com_dro.get(coluna) for coluna in BASE_COLUNAS])

    evento_valido = dict(CAMPOS_EVENTO_PADRAO)
    evento_valido.update(
        idEvento="EVT2",
        categoriaNivel1="2",
        categoriaNivel2="21",
        codigoEventoOrigem="COD2",
        valorPerdaEfetiva="10.00",
    )
    aba_base.append([evento_valido.get(coluna) for coluna in BASE_COLUNAS])

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(
        [CABECALHO_VALIDO.get(coluna) for coluna in CABECALHO_COLUNAS]
    )
    caminho = tmp_path / "erro_oficial.xlsx"
    workbook.save(caminho)

    eventos_consolidados: list[str] = []
    consolidar_original = conversion.consolidar_eventos

    def registrar_consolidacao(eventos, data_base):
        eventos_consolidados.extend(eventos)
        return consolidar_original(eventos, data_base)

    monkeypatch.setattr(conversion, "consolidar_eventos", registrar_consolidacao)
    resultado = conversion.processar(caminho, tmp_path / "saida")

    assert any(o.codigo == "DRO001301" for o in resultado.ocorrencias)
    assert eventos_consolidados == ["EVT2"]
