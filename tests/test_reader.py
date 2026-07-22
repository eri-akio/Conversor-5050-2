"""Testes da Fase 2: leitura e validacao estrutural (src/reader.py)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.reader import (
    BASE_COLUNAS,
    CABECALHO_COLUNAS,
    ArquivoInvalido,
    PlanilhaInvalida,
    extrair_cabecalho,
    ler_planilha,
)


def _valores_linha_base() -> tuple[object, ...]:
    return tuple(f"valor-{coluna}" for coluna in BASE_COLUNAS)


def _valores_linha_cabecalho() -> tuple[object, ...]:
    return tuple(f"valor-{coluna}" for coluna in CABECALHO_COLUNAS)


def _construir_workbook(
    *,
    colunas_base: tuple[str, ...] = BASE_COLUNAS,
    colunas_cabecalho: tuple[str, ...] = CABECALHO_COLUNAS,
    linhas_base: list[tuple[object, ...]] | None = None,
    linhas_cabecalho: list[tuple[object, ...]] | None = None,
) -> Workbook:
    workbook = Workbook()
    aba_base = workbook.active
    aba_base.title = "Base"
    aba_base.append(list(colunas_base))
    for linha in (
        linhas_base if linhas_base is not None else [_valores_linha_base()]
    ):
        aba_base.append(list(linha))

    aba_cabecalho = workbook.create_sheet("Cabecalho")
    aba_cabecalho.append(list(colunas_cabecalho))
    for linha in (
        linhas_cabecalho
        if linhas_cabecalho is not None
        else [_valores_linha_cabecalho()]
    ):
        aba_cabecalho.append(list(linha))

    return workbook


def _salvar(workbook: Workbook, tmp_path: Path, nome: str = "planilha.xlsx") -> Path:
    caminho = tmp_path / nome
    workbook.save(caminho)
    return caminho


def test_planilha_valida_e_lida_com_sucesso(tmp_path: Path) -> None:
    caminho = _salvar(_construir_workbook(), tmp_path)

    planilha = ler_planilha(caminho)

    assert planilha.cabecalhos_base == BASE_COLUNAS
    assert planilha.cabecalhos_cabecalho == CABECALHO_COLUNAS
    assert len(planilha.linhas_base()) == 1
    assert planilha.linha_cabecalho() is not None


def test_extrair_cabecalho_normaliza_formatacao(tmp_path: Path) -> None:
    linha_cabecalho = {
        "codigoDocumento": "5050",
        "dataBase": "2026-06",
        "codigoConglomerado": "c0099999",
        "cnpj": "46.169.337/0001-28",
        "tipoRemessa": "i",
        "opcaoPorProvisaoAcumulada": "n",
    }
    valores = tuple(linha_cabecalho[coluna] for coluna in CABECALHO_COLUNAS)
    caminho = _salvar(
        _construir_workbook(linhas_cabecalho=[valores]), tmp_path
    )

    planilha = ler_planilha(caminho)
    cabecalho = extrair_cabecalho(planilha)

    assert cabecalho["codigoConglomerado"].valor == "C0099999"
    assert cabecalho["cnpj"].valor == "46169337"
    assert cabecalho["tipoRemessa"].valor == "I"
    assert cabecalho["opcaoPorProvisaoAcumulada"].valor == "N"


def test_arquivo_ausente_gera_falha_tecnica(tmp_path: Path) -> None:
    with pytest.raises(ArquivoInvalido):
        ler_planilha(tmp_path / "nao_existe.xlsx")


def test_extensao_invalida_gera_falha_tecnica(tmp_path: Path) -> None:
    caminho = tmp_path / "planilha.csv"
    caminho.write_text("idEvento\n1\n", encoding="utf-8")

    with pytest.raises(ArquivoInvalido):
        ler_planilha(caminho)


def test_aba_ausente_gera_xlsx_aba_001(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Base"
    workbook.active.append(list(BASE_COLUNAS))
    workbook.active.append(list(_valores_linha_base()))
    # Sem a aba Cabecalho.
    caminho = _salvar(workbook, tmp_path)

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-ABA-001"


def test_nomes_de_aba_sao_case_insensitive_e_com_trim(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = " base "
    workbook.active.append(list(BASE_COLUNAS))
    workbook.active.append(list(_valores_linha_base()))
    aba_cabecalho = workbook.create_sheet("CABECALHO")
    aba_cabecalho.append(list(CABECALHO_COLUNAS))
    aba_cabecalho.append(list(_valores_linha_cabecalho()))
    caminho = _salvar(workbook, tmp_path)

    planilha = ler_planilha(caminho)

    assert len(planilha.linhas_base()) == 1


def test_coluna_obrigatoria_ausente_gera_xlsx_col_001(tmp_path: Path) -> None:
    colunas_incompletas = tuple(c for c in BASE_COLUNAS if c != "idEvento")
    valores = tuple(f"valor-{c}" for c in colunas_incompletas)
    caminho = _salvar(
        _construir_workbook(
            colunas_base=colunas_incompletas,
            linhas_base=[valores],
        ),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-COL-001"
    assert "idEvento" in excinfo.value.ocorrencia.campos


def test_cabecalho_duplicado_gera_xlsx_col_002(tmp_path: Path) -> None:
    colunas_duplicadas = BASE_COLUNAS + ("idEvento",)
    valores = tuple(f"valor-{c}" for c in colunas_duplicadas)
    caminho = _salvar(
        _construir_workbook(
            colunas_base=colunas_duplicadas,
            linhas_base=[valores],
        ),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-COL-002"


def test_alias_e_nome_canonico_juntos_geram_ambiguidade(tmp_path: Path) -> None:
    colunas = BASE_COLUNAS + ("ligacaoRiscoSocioambiental",)
    valores = tuple(f"valor-{c}" for c in colunas)
    caminho = _salvar(
        _construir_workbook(colunas_base=colunas, linhas_base=[valores]),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-COL-002"


def test_alias_sozinho_e_aceito_no_lugar_do_nome_canonico(tmp_path: Path) -> None:
    colunas = tuple(
        "ligacaoRiscoSocioambiental" if c == "ligadoRiscoSocioAmbiental" else c
        for c in BASE_COLUNAS
    )
    valores = tuple(f"valor-{c}" for c in colunas)
    caminho = _salvar(
        _construir_workbook(colunas_base=colunas, linhas_base=[valores]),
        tmp_path,
    )

    planilha = ler_planilha(caminho)

    assert "ligacaoRiscoSocioambiental" in planilha.cabecalhos_base


def test_base_sem_linha_de_dados_gera_xlsx_base_001(tmp_path: Path) -> None:
    caminho = _salvar(
        _construir_workbook(linhas_base=[]),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-BASE-001"


def test_cabecalho_sem_linha_de_dados_gera_xlsx_cab_002(tmp_path: Path) -> None:
    caminho = _salvar(
        _construir_workbook(linhas_cabecalho=[]),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-CAB-002"


def test_cabecalho_com_mais_de_uma_linha_gera_xlsx_cab_002(tmp_path: Path) -> None:
    caminho = _salvar(
        _construir_workbook(
            linhas_cabecalho=[
                _valores_linha_cabecalho(),
                _valores_linha_cabecalho(),
            ]
        ),
        tmp_path,
    )

    with pytest.raises(PlanilhaInvalida) as excinfo:
        ler_planilha(caminho)

    assert excinfo.value.ocorrencia.codigo == "XLSX-CAB-002"
